#Need to change for each user
path_to_onedrive = "/Users/nboyer/OneDrive - University of North Carolina at Chapel Hill"


path = path_to_onedrive+"/Documents - UNC ITS RC Engage Support/Application Form to Request Registration as UNC-CH Affiliate 1.xlsx"

from openpyxl import load_workbook

#Selecting specific sheet
wb = load_workbook(path)
sheet = wb.worksheets[0] # 0 1 2 3 or any

#reading cell
#chr(65)
ind = 2
while str(sheet.cell(row=ind, column=1).value) != "None":
    print(sheet.cell(row=ind, column=1).value)
    #print(str(sheet.cell(row=ind, column=1).value) != "None")
    if sheet.cell(row=ind, column=23).value != "Sent to Unsigned":
        
        from PyPDF2 import PdfWriter, PdfReader
        import io
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter

        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=letter)
        
        re = []
        for j in range(22):
            rd = str(sheet[chr(65+j)+str(ind)].value)
            if rd != "None":
                re.append(rd)
        print(re)
        can.drawString(120, 690, re[4])
        can.drawString(270, 690, re[5])
        can.drawString(400, 690, re[6])
        can.drawString(150, 560, re[7])
        can.drawString(250, 580, re[8])
        can.drawString(450, 580, re[9])
        can.drawString(150, 620, re[10] + " " + re[11] + " " + re[12] + " " + re[13])
        can.drawString(200, 600, re[14] + " " + re[15] + " " + re[16] + " " + re[17])
        st =re[18][:-9]
        can.drawString(470, 660, st[5:7] + "/" + st[8:] + "/" +st[:4])
        if (re[19] == "Male"):
            can.drawString(200, 640, "X")
        elif (re[19] == "Female"):
            can.drawString(245, 640, "X")

        if len(re) > 20:
            can.drawString(410, 730, re[20])

        #our info
        '''
        start date (7th of the current month [or next month if today is the last day of the month])
                    
        end date (1 year from the start date; day should be the 7th)
                    
        John McGee's PID: 711859032
                    
        today's date next to John's signature line
        '''

        from datetime import datetime
        #print(datetime.today().strftime('%m/7/%Y'))
        abc = int(datetime.today().strftime('%Y'))+1
        #print(datetime.today().strftime('%m/7/' + str(abc)))
        #print(datetime.today().strftime('%m/%d/%Y'))
        #print(datetime.today())
        can.drawString(230, 370, datetime.today().strftime('%m/7/%Y'))
        can.drawString(450, 370, datetime.today().strftime('%m/7/' + str(abc)))
        can.drawString(470, 170, datetime.today().strftime('%m/%d/%Y'))
        #711859032
        JM = "711859032"
        can.drawString(120, 295,JM)
        can.save()

        #move to the beginning of the StringIO buffer
        packet.seek(0)

        # create a new PDF with Reportlab
        new_pdf = PdfReader(packet)
        # read your existing PDF
        existing_pdf = PdfReader(open("form.pdf", "rb"))
        output = PdfWriter()
        # add the "watermark" (which is the new pdf) on the existing page
        page = existing_pdf.pages[0]
        page.merge_page(new_pdf.pages[0])
        output.add_page(page)
        # finally, write "output" to a real file
        paths = path_to_onedrive+"/Unsigned/"+re[6]+"_" + re[4]+"_"+str(datetime.today().strftime('%m_%d_%Y'))+"-affiliate-application_rc_sponsored.pdf"

        output_stream = open(paths, "wb")
        output.write(output_stream)
        output_stream.close()
        sheet.cell(row=ind, column=23).value="Sent to Unsigned"
        wb.save(path)
    else:
        print("Done")
    ind +=1